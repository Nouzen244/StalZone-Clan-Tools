"""
Модуль посещаемости v3.2
КВ работает ТОЛЬКО по расписанию
Личная статистика для всех
"""

import discord
from discord.ext import commands, tasks
import logging
from datetime import datetime, timedelta
from typing import Optional, List, Dict
from pathlib import Path
from ranks import RANK_ORDER, RANK_NAMES, RANK_ICONS

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger('Attendance')


def format_date(date_obj=None, date_str=None) -> str:
    if date_obj:
        return date_obj.strftime('%d-%m-%Y')
    if date_str:
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            return dt.strftime('%d-%m-%Y')
        except:
            return date_str
    return datetime.now().strftime('%d-%m-%Y')

def parse_date(date_str: str) -> Optional[datetime]:
    for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def date_for_db(date_obj=None) -> str:
    if date_obj:
        return date_obj.strftime('%Y-%m-%d')
    return datetime.now().strftime('%Y-%m-%d')

def format_duration(seconds: int) -> str:
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    if hours > 0:
        return f"{hours}ч {minutes}м"
    return f"{minutes}м"


def officer_or_higher():
    async def predicate(ctx: commands.Context) -> bool:
        if ctx.author.guild_permissions.administrator:
            return True
        return await ctx.bot.has_permission(ctx.author, 'officer')
    return commands.check(predicate)


# ============================================
# VIEWS ДЛЯ РЕДАКТИРОВАНИЯ КВ
# ============================================

class KVMemberSelect(discord.ui.Select):
    """Select Menu для выбора отсутствующего игрока"""
    
    def __init__(self, bot, absent_members: List, date: str, guild_id: int):
        self.bot = bot
        self.date = date
        self.guild_id = guild_id
        
        options = []
        for member, data, role_type in absent_members[:25]:
            excused = data.get('excused')
            label = member.display_name[:100]
            emoji = "📝" if excused == 'У/П' else "❌"
            
            options.append(discord.SelectOption(
                label=label,
                value=str(member.id),
                description=f"Роль: {role_type}" + (" [У/П]" if excused else ""),
                emoji=emoji
            ))
        
        super().__init__(
            placeholder="📋 Выберите игрока для обработки...",
            options=options if options else [discord.SelectOption(label="Нет данных", value="none")],
            min_values=1,
            max_values=1
        )
    
    async def callback(self, interaction: discord.Interaction):
        if self.values[0] == "none":
            return
        
        member_id = int(self.values[0])
        member = interaction.guild.get_member(member_id)
        
        if not member:
            await interaction.response.send_message("❌ Пользователь не найден!", ephemeral=True)
            return
        
        embed = discord.Embed(
            title=f"📝 Обработка: {member.display_name}",
            description="Пользователь **отсутствовал** на КВ?",
            color=discord.Color.orange()
        )
        embed.set_thumbnail(url=member.display_avatar.url)
        
        view = KVStep1View(self.bot, member, self.date, self.guild_id)
        await interaction.response.send_message(embed=embed, view=view, ephemeral=True)


class KVStep1View(discord.ui.View):
    """Шаг 1: Подтверждение отсутствия"""
    
    def __init__(self, bot, member: discord.Member, date: str, guild_id: int):
        super().__init__(timeout=300)
        self.bot = bot
        self.member = member
        self.date = date
        self.guild_id = guild_id
    
    @discord.ui.button(label="Да, отсутствовал", style=discord.ButtonStyle.danger, emoji="❌")
    async def confirm_absent(self, interaction: discord.Interaction, button):
        embed = discord.Embed(
            title=f"📝 {self.member.display_name}",
            description="Причина **уважительная**?",
            color=discord.Color.orange()
        )
        embed.set_thumbnail(url=self.member.display_avatar.url)
        
        view = KVStep2View(self.bot, self.member, self.date, self.guild_id)
        await interaction.response.edit_message(embed=embed, view=view)
    
    @discord.ui.button(label="Нет, был на КВ", style=discord.ButtonStyle.success, emoji="✅")
    async def mark_present(self, interaction: discord.Interaction, button):
        await self.bot.db.execute('''
            UPDATE kv_attendance 
            SET present = 1, excused = NULL, reason = 'Исправлено вручную', processed_by = ?
            WHERE guild_id = ? AND date = ? AND user_id = ?
        ''', (interaction.user.id, self.guild_id, self.date, self.member.id))
        await self.bot.db.commit()
        
        embed = discord.Embed(
            title="✅ Исправлено!",
            description=f"**{self.member.display_name}** отмечен как **присутствующий**",
            color=discord.Color.green()
        )
        embed.add_field(name="📅 Дата", value=format_date(date_str=self.date), inline=True)
        embed.add_field(name="👤 Изменил", value=interaction.user.mention, inline=True)
        
        await interaction.response.edit_message(embed=embed, view=None)


class KVStep2View(discord.ui.View):
    """Шаг 2: Уважительная причина?"""
    
    def __init__(self, bot, member: discord.Member, date: str, guild_id: int):
        super().__init__(timeout=300)
        self.bot = bot
        self.member = member
        self.date = date
        self.guild_id = guild_id
    
    @discord.ui.button(label="Да, уважительная", style=discord.ButtonStyle.success, emoji="📝")
    async def excused_yes(self, interaction: discord.Interaction, button):
        embed = discord.Embed(
            title=f"📝 {self.member.display_name}",
            description="Выберите **причину** отсутствия:",
            color=discord.Color.blue()
        )
        embed.set_thumbnail(url=self.member.display_avatar.url)
        
        view = KVStep3View(self.bot, self.member, self.date, self.guild_id)
        await interaction.response.edit_message(embed=embed, view=view)
    
    @discord.ui.button(label="Нет, неуважительная", style=discord.ButtonStyle.danger, emoji="⛔")
    async def excused_no(self, interaction: discord.Interaction, button):
        await self.bot.db.execute('''
            UPDATE kv_attendance 
            SET present = 0, excused = NULL, reason = 'Неуважительный пропуск', processed_by = ?
            WHERE guild_id = ? AND date = ? AND user_id = ?
        ''', (interaction.user.id, self.guild_id, self.date, self.member.id))
        await self.bot.db.commit()
        
        embed = discord.Embed(
            title="⛔ Неуважительный пропуск",
            description=f"**{self.member.display_name}** — пропуск без уважительной причины",
            color=discord.Color.red()
        )
        embed.add_field(name="📅 Дата", value=format_date(date_str=self.date), inline=True)
        embed.add_field(name="👤 Обработал", value=interaction.user.mention, inline=True)
        
        await interaction.response.edit_message(embed=embed, view=None)


class KVStep3View(discord.ui.View):
    """Шаг 3: Выбор причины"""
    
    def __init__(self, bot, member: discord.Member, date: str, guild_id: int):
        super().__init__(timeout=300)
        self.bot = bot
        self.member = member
        self.date = date
        self.guild_id = guild_id
    
    @discord.ui.select(
        placeholder="📋 Выберите причину...",
        options=[
            discord.SelectOption(label="Болезнь", value="Болезнь", emoji="🏥"),
            discord.SelectOption(label="Работа", value="Работа", emoji="💼"),
            discord.SelectOption(label="Семья", value="Семья", emoji="👨‍👩‍👧"),
            discord.SelectOption(label="Технические проблемы", value="Тех.проблемы", emoji="🔧"),
            discord.SelectOption(label="Учёба", value="Учёба", emoji="📚"),
            discord.SelectOption(label="В отъезде", value="Отъезд", emoji="✈️"),
            discord.SelectOption(label="Играл оффлайн", value="Оффлайн", emoji="🎮"),
            discord.SelectOption(label="Другое", value="Другое", emoji="📝"),
        ]
    )
    async def select_reason(self, interaction: discord.Interaction, select):
        reason = select.values[0]
        
        await self.bot.db.execute('''
            UPDATE kv_attendance 
            SET present = 0, excused = 'У/П', reason = ?, processed_by = ?
            WHERE guild_id = ? AND date = ? AND user_id = ?
        ''', (reason, interaction.user.id, self.guild_id, self.date, self.member.id))
        await self.bot.db.commit()
        
        embed = discord.Embed(
            title="📝 У/П — Уважительная причина",
            description=f"**{self.member.display_name}**",
            color=discord.Color.blue()
        )
        embed.add_field(name="📋 Причина", value=reason, inline=True)
        embed.add_field(name="📅 Дата", value=format_date(date_str=self.date), inline=True)
        embed.add_field(name="👤 Обработал", value=interaction.user.mention, inline=True)
        
        await interaction.response.edit_message(embed=embed, view=None)
    
    @discord.ui.button(label="⏭️ Пропустить", style=discord.ButtonStyle.secondary, row=1)
    async def skip(self, interaction: discord.Interaction, button):
        await self.bot.db.execute('''
            UPDATE kv_attendance 
            SET present = 0, excused = 'У/П', reason = 'Не указана', processed_by = ?
            WHERE guild_id = ? AND date = ? AND user_id = ?
        ''', (interaction.user.id, self.guild_id, self.date, self.member.id))
        await self.bot.db.commit()
        
        embed = discord.Embed(
            title="📝 У/П — Причина не указана",
            description=f"**{self.member.display_name}** отмечен с уважительной причиной",
            color=discord.Color.blue()
        )
        
        await interaction.response.edit_message(embed=embed, view=None)


class SundayChoiceButton(discord.ui.Button):
    """Кнопка выбора воскресного события (Потасовка / Захват базы)"""

    EMOJIS = {'Потасовка': '🥊', 'Захват базы': '🏴'}

    def __init__(self, bot, date: str, guild_id: int, event_name: str, is_live: bool, current_choice):
        chosen = (current_choice == event_name)
        super().__init__(
            label=event_name,
            style=discord.ButtonStyle.success if chosen else discord.ButtonStyle.secondary,
            emoji="✅" if chosen else self.EMOJIS.get(event_name, "⚔️"),
            row=2
        )
        self.bot = bot
        self.date = date
        self.guild_id = guild_id
        self.event_name = event_name
        self.is_live = is_live

    async def callback(self, interaction: discord.Interaction):
        # Выбирать может только офицер+ или админ
        is_admin = interaction.user.guild_permissions.administrator
        if not is_admin and not await self.bot.has_permission(interaction.user, 'officer'):
            await interaction.response.send_message(
                "🚫 Только офицеры могут выбирать воскресное событие!", ephemeral=True
            )
            return

        await self.bot.set_sunday_choice(self.guild_id, self.date, self.event_name, interaction.user.id)

        # Подсвечиваем выбранную кнопку
        for item in self.view.children:
            if isinstance(item, SundayChoiceButton):
                chosen = item.event_name == self.event_name
                item.style = discord.ButtonStyle.success if chosen else discord.ButtonStyle.secondary
                item.emoji = "✅" if chosen else SundayChoiceButton.EMOJIS.get(item.event_name, "⚔️")

        # Обновляем заголовок отчёта
        embed = interaction.message.embeds[0] if interaction.message.embeds else None
        if embed:
            embed.title = (f"⚔️ КВ: {self.event_name} (LIVE)"
                           if self.is_live else f"⚔️ Отчёт КВ: {self.event_name}")

        await interaction.response.edit_message(embed=embed, view=self.view)


class KVReportView(discord.ui.View):
    """View для отчёта КВ с кнопками редактирования"""

    def __init__(self, bot, absent_members: List, date: str, guild_id: int, is_live: bool = False,
                 sunday: bool = False, sunday_choices: List = None):
        super().__init__(timeout=600)
        self.bot = bot
        self.absent = absent_members
        self.date = date
        self.guild_id = guild_id
        self.is_live = is_live

        # Добавляем Select Menu если есть отсутствующие
        if absent_members:
            self.add_item(KVMemberSelect(bot, absent_members, date, guild_id))

        # Кнопки выбора воскресного события (Потасовка / Захват базы)
        if sunday and sunday_choices:
            current_choice = bot.sunday_choices.get(f"{guild_id}:{date}")
            for event_name in sunday_choices:
                self.add_item(SundayChoiceButton(bot, date, guild_id, event_name, is_live, current_choice))

    @discord.ui.button(label="🔄 Обновить", style=discord.ButtonStyle.secondary, row=1)
    async def refresh(self, interaction: discord.Interaction, button):
        await interaction.response.defer()
        # Вызываем команду !kv заново
        ctx = await self.bot.get_context(interaction.message)
        if self.is_live:
            await ctx.invoke(self.bot.get_command('kv'))
        else:
            await ctx.invoke(self.bot.get_command('kv'), date=format_date(date_str=self.date))


class KVEditView(discord.ui.View):
    """View для редактирования КВ"""
    
    def __init__(self, bot, date: str, guild_id: int):
        super().__init__(timeout=300)
        self.bot = bot
        self.date = date
        self.guild_id = guild_id
    
    @discord.ui.button(label="📝 Редактировать участника", style=discord.ButtonStyle.primary)
    async def edit_member(self, interaction: discord.Interaction, button):
        modal = KVEditModal(self.bot, self.date, self.guild_id)
        await interaction.response.send_modal(modal)
    
    @discord.ui.button(label="✅ Отметить всех присутствующими", style=discord.ButtonStyle.success)
    async def mark_all_present(self, interaction: discord.Interaction, button):
        await self.bot.db.execute('''
            UPDATE kv_attendance SET present = 1, processed_by = ?
            WHERE guild_id = ? AND date = ?
        ''', (interaction.user.id, self.guild_id, self.date))
        await self.bot.db.commit()
        
        await interaction.response.send_message("✅ Все участники отмечены как присутствующие!", ephemeral=True)


class KVEditModal(discord.ui.Modal, title="📝 Редактировать участника"):
    """Модальное окно для поиска и редактирования"""
    
    username = discord.ui.TextInput(
        label="Никнейм или ID участника",
        placeholder="Введите ник или Discord ID...",
        required=True,
        max_length=100
    )
    
    def __init__(self, bot, date: str, guild_id: int):
        super().__init__()
        self.bot = bot
        self.date = date
        self.guild_id = guild_id
    
    async def on_submit(self, interaction: discord.Interaction):
        search = self.username.value.strip()
        
        # Поиск участника
        member = None
        
        if search.isdigit():
            member = interaction.guild.get_member(int(search))
        
        if not member:
            search_lower = search.lower()
            for m in interaction.guild.members:
                if search_lower in m.display_name.lower() or search_lower in m.name.lower():
                    member = m
                    break
        
        if not member:
            await interaction.response.send_message(f"❌ Участник **{search}** не найден!", ephemeral=True)
            return
        
        # Показываем меню редактирования
        embed = discord.Embed(
            title=f"📝 Редактирование: {member.display_name}",
            description="Выберите новый статус:",
            color=discord.Color.blue()
        )
        embed.set_thumbnail(url=member.display_avatar.url)
        
        view = KVDirectEditView(self.bot, member, self.date, self.guild_id)
        await interaction.response.send_message(embed=embed, view=view, ephemeral=True)


class KVDirectEditView(discord.ui.View):
    """Прямое редактирование статуса"""
    
    def __init__(self, bot, member: discord.Member, date: str, guild_id: int):
        super().__init__(timeout=300)
        self.bot = bot
        self.member = member
        self.date = date
        self.guild_id = guild_id
    
    @discord.ui.button(label="✅ Присутствовал", style=discord.ButtonStyle.success)
    async def mark_present(self, interaction: discord.Interaction, button):
        await self.bot.db.execute('''
            UPDATE kv_attendance 
            SET present = 1, excused = NULL, reason = 'Изменено вручную', processed_by = ?
            WHERE guild_id = ? AND date = ? AND user_id = ?
        ''', (interaction.user.id, self.guild_id, self.date, self.member.id))
        await self.bot.db.commit()
        
        await interaction.response.edit_message(
            content=f"✅ **{self.member.display_name}** → Присутствовал",
            embed=None, view=None
        )
    
    @discord.ui.button(label="❌ Отсутствовал", style=discord.ButtonStyle.danger)
    async def mark_absent(self, interaction: discord.Interaction, button):
        await self.bot.db.execute('''
            UPDATE kv_attendance 
            SET present = 0, excused = NULL, reason = NULL, processed_by = ?
            WHERE guild_id = ? AND date = ? AND user_id = ?
        ''', (interaction.user.id, self.guild_id, self.date, self.member.id))
        await self.bot.db.commit()
        
        await interaction.response.edit_message(
            content=f"❌ **{self.member.display_name}** → Отсутствовал",
            embed=None, view=None
        )
    
    @discord.ui.button(label="📝 У/П (уважительная)", style=discord.ButtonStyle.primary)
    async def mark_excused(self, interaction: discord.Interaction, button):
        embed = discord.Embed(
            title=f"📝 {self.member.display_name}",
            description="Выберите причину:",
            color=discord.Color.blue()
        )
        view = KVStep3View(self.bot, self.member, self.date, self.guild_id)
        await interaction.response.edit_message(embed=embed, view=view)


# ============================================
# ОСНОВНОЙ КОГ
# ============================================

class AttendanceCog(commands.Cog, name="Посещаемость"):
    """Команды КВ и отчётности"""

    def __init__(self, bot):
        self.bot = bot
        logger.info("📊 AttendanceCog v3.2 загружен")

    # ========================================
    # КВ - ТОЛЬКО ПО РАСПИСАНИЮ
    # ========================================
    
    @commands.command(name='kv', aliases=['кв'])
    @officer_or_higher()
    async def kv_report(self, ctx: commands.Context, date: str = None):
        """
        Отчёт по КВ
        !kv - текущее КВ (только если идёт по расписанию)
        !kv 17-01-2026 - отчёт за дату
        """
        guild = ctx.guild
        guild_id = guild.id
        settings = self.bot.guild_settings.get(guild_id, {})
        
        # Проверяем настройки
        kv_vc_id = settings.get('kv_vc_channel_id')
        if not kv_vc_id:
            await ctx.send("❌ Канал КВ не настроен!\n`!setvc kv #канал`")
            return
        
        kv_vc = guild.get_channel(kv_vc_id)
        if not kv_vc:
            await ctx.send("❌ Канал КВ не найден!")
            return
        
        now = datetime.now(self.bot.timezone)
        
        # Без даты = текущее КВ
        if not date:
            current_kv = self.bot.get_current_kv_schedule(guild_id)
            
            if not current_kv:
                # Расписание фиксированное — показываем сегодняшнее событие
                day_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
                today_weekday = now.weekday()
                today_str = date_for_db(now)
                today_event = None

                for sched in self.bot.get_guild_schedules(guild_id):
                    if today_weekday in sched['days_of_week']:
                        name = self.bot.get_event_name(guild_id, today_str, sched)
                        today_event = f"**{name}** в {sched['start_time']}-{sched['end_time']}"
                        break

                embed = discord.Embed(
                    title="⚠️ Сейчас нет активного КВ",
                    description=(f"**Сегодня ({day_names[today_weekday]}):** {today_event}\n\n" if today_event else "")
                               + "Используйте `!kv DD-MM-YYYY` для просмотра отчёта за дату",
                    color=discord.Color.orange()
                )
                await ctx.send(embed=embed)
                return
            
            # КВ сейчас идёт - показываем live статус
            await self._show_live_kv(ctx, current_kv, kv_vc, now)
        
        else:
            # С датой = отчёт за прошедшее КВ
            target_date = parse_date(date)
            if not target_date:
                await ctx.send("❌ Неверный формат даты! Используйте: `DD-MM-YYYY`")
                return
            
            await self._show_kv_report(ctx, target_date, kv_vc)
    
    async def _show_live_kv(self, ctx, schedule: dict, kv_vc: discord.VoiceChannel, now: datetime):
        """Показывает live статус текущего КВ"""
        guild = ctx.guild
        guild_id = guild.id
        date_str = date_for_db(now)
        
        # Получаем всех участников сервера (не ботов)
        all_members = [m for m in guild.members if not m.bot]
        
        # Кто сейчас в VC
        members_in_vc = {m.id for m in kv_vc.members if not m.bot}
        
        # Создаём/обновляем записи в БД
        for member in all_members:
            is_present = member.id in members_in_vc
            role_type = self.bot.get_member_role_type(member)
            
            await self.bot.db.execute('''
                INSERT INTO kv_attendance 
                (guild_id, schedule_id, date, kv_time, user_id, discord_name, role_type, present)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(guild_id, date, user_id, schedule_id) DO UPDATE SET
                    present = MAX(kv_attendance.present, excluded.present),
                    role_type = excluded.role_type
            ''', (
                guild_id, schedule['id'], date_str,
                f"{schedule['start_time']}-{schedule['end_time']}",
                member.id, member.display_name, role_type,
                1 if is_present else 0
            ))
        
        await self.bot.db.commit()
        
        # Формируем отчёт
        present = []
        absent = []
        
        role_icons = RANK_ICONS
        
        for member in all_members:
            role_type = self.bot.get_member_role_type(member)
            role_icon = role_icons.get(role_type, '')
            
            if member.id in members_in_vc:
                # Получаем время в VC
                session_key = f"{guild_id}:{member.id}"
                session = self.bot.all_voice_sessions.get(session_key)
                if session and session.get('channel_id') == kv_vc.id:
                    duration = (now - session['join_time']).total_seconds()
                    present.append((member, role_icon, role_type, int(duration)))
                else:
                    present.append((member, role_icon, role_type, 0))
            else:
                absent.append((member, role_icon, role_type))
        
        # Сортировка
        present.sort(key=lambda x: (-x[3], x[0].display_name.lower()))
        absent.sort(key=lambda x: x[0].display_name.lower())
        
        # Embed
        event_name = self.bot.get_event_name(guild_id, date_str, schedule)
        embed = discord.Embed(
            title=f"⚔️ КВ: {event_name} (LIVE)",
            description=f"🕐 **Время:** {schedule['start_time']} - {schedule['end_time']}\n"
                       f"🎤 **Канал:** {kv_vc.mention}",
            color=discord.Color.red(),
            timestamp=now
        )
        
        # Присутствующие
        if present:
            lines = []
            for member, icon, role_type, duration in present[:15]:
                time_str = f" ({format_duration(duration)})" if duration > 0 else ""
                lines.append(f"🟢{icon} {member.display_name}{time_str}")
            
            if len(present) > 15:
                lines.append(f"... и ещё {len(present) - 15}")
            
            embed.add_field(
                name=f"✅ В канале ({len(present)})",
                value="\n".join(lines) or "—",
                inline=True
            )
        else:
            embed.add_field(name="✅ В канале (0)", value="Пусто", inline=True)
        
        # Отсутствующие
        if absent:
            lines = []
            for member, icon, role_type in absent[:15]:
                lines.append(f"❌{icon} {member.display_name}")
            
            if len(absent) > 15:
                lines.append(f"... и ещё {len(absent) - 15}")
            
            embed.add_field(
                name=f"❌ Отсутствуют ({len(absent)})",
                value="\n".join(lines) or "—",
                inline=True
            )
        
        # Статистика
        total = len(all_members)
        present_pct = (len(present) / total * 100) if total > 0 else 0
        
        embed.add_field(
            name="📊 Статистика",
            value=f"**Всего:** {total}\n"
                  f"**Присутствует:** {len(present)} ({present_pct:.0f}%)\n"
                  f"**Отсутствует:** {len(absent)}",
            inline=False
        )
        
        embed.set_footer(text=f"Обновлено • {ctx.author.display_name}")
        
        # View с кнопками
        absent_data = [(m, {}, rt) for m, _, rt in absent]
        sunday = bool(schedule.get('choices'))
        view = KVReportView(self.bot, absent_data, date_str, guild_id, is_live=True,
                            sunday=sunday, sunday_choices=schedule.get('choices'))

        await ctx.send(embed=embed, view=view)
    
    async def _show_kv_report(self, ctx, target_date: datetime, kv_vc: discord.VoiceChannel):
        """Показывает отчёт за прошедшую дату"""
        guild = ctx.guild
        guild_id = guild.id
        date_str = date_for_db(target_date)
        
        # Проверяем было ли КВ в эту дату
        schedules = self.bot.get_guild_schedules(guild_id)
        day_of_week = target_date.weekday()
        
        matching_schedule = None
        for sched in schedules:
            if day_of_week in sched['days_of_week']:
                matching_schedule = sched
                break
        
        if not matching_schedule:
            day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
            await ctx.send(f"⚠️ На **{format_date(target_date)}** ({day_names[day_of_week]}) не было КВ по расписанию!")
            return
        
        kv_time = f"{matching_schedule['start_time']}-{matching_schedule['end_time']}"
        
        # Получаем данные из БД
        attendance_data = {}
        async with self.bot.db.execute('''
            SELECT user_id, present, excused, reason, vc_time_seconds, role_type
            FROM kv_attendance
            WHERE guild_id = ? AND date = ?
        ''', (guild_id, date_str)) as cursor:
            async for row in cursor:
                attendance_data[row[0]] = {
                    'present': row[1],
                    'excused': row[2],
                    'reason': row[3],
                    'vc_time': row[4] or 0,
                    'role_type': row[5]
                }
        
        # Если нет данных - синхронизируем с сессиями
        if not attendance_data:
            # Получаем всех кто был в VC КВ в эту дату
            async with self.bot.db.execute('''
                SELECT user_id, SUM(duration_seconds) as total_time
                FROM voice_sessions
                WHERE guild_id = ? AND date = ? AND channel_id = ?
                GROUP BY user_id
            ''', (guild_id, date_str, kv_vc.id)) as cursor:
                sessions = await cursor.fetchall()
            
            vc_users = {row[0]: row[1] for row in sessions}
            
            # Добавляем всех участников сервера
            for member in guild.members:
                if member.bot:
                    continue
                
                role_type = self.bot.get_member_role_type(member)
                was_present = member.id in vc_users and vc_users[member.id] > 60
                vc_time = vc_users.get(member.id, 0)
                
                await self.bot.db.execute('''
                    INSERT OR IGNORE INTO kv_attendance 
                    (guild_id, schedule_id, date, kv_time, user_id, discord_name, role_type, present, vc_time_seconds)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    guild_id, matching_schedule['id'], date_str, kv_time,
                    member.id, member.display_name, role_type,
                    1 if was_present else 0, int(vc_time)
                ))
                
                attendance_data[member.id] = {
                    'present': 1 if was_present else 0,
                    'excused': None,
                    'reason': None,
                    'vc_time': int(vc_time),
                    'role_type': role_type
                }
            
            await self.bot.db.commit()
        
        # Формируем списки
        present = []
        absent = []
        
        role_icons = RANK_ICONS
        
        for member in guild.members:
            if member.bot:
                continue
            
            data = attendance_data.get(member.id, {})
            role_type = self.bot.get_member_role_type(member)
            role_icon = role_icons.get(role_type, '')
            
            if data.get('present'):
                vc_time = data.get('vc_time', 0)
                present.append((member, role_icon, role_type, vc_time))
            else:
                absent.append((member, data, role_type))
        
        # Сортировка
        present.sort(key=lambda x: (-x[3], x[0].display_name.lower()))
        absent.sort(key=lambda x: x[0].display_name.lower())
        
        # Embed
        event_name = self.bot.get_event_name(guild_id, date_str, matching_schedule)
        embed = discord.Embed(
            title=f"⚔️ Отчёт КВ: {event_name}",
            description=f"📅 **Дата:** {format_date(target_date)}\n"
                       f"🕐 **Время:** {kv_time}\n"
                       f"🎤 **Канал:** {kv_vc.mention}",
            color=discord.Color.blue(),
            timestamp=datetime.now()
        )
        
        # Присутствовали
        if present:
            lines = []
            for member, icon, role_type, vc_time in present[:15]:
                time_str = f" ({format_duration(vc_time)})" if vc_time > 0 else ""
                lines.append(f"✅{icon} {member.display_name}{time_str}")
            
            if len(present) > 15:
                lines.append(f"... и ещё {len(present) - 15}")
            
            embed.add_field(
                name=f"✅ Присутствовали ({len(present)})",
                value="\n".join(lines) or "—",
                inline=True
            )
        
        # Отсутствовали
        if absent:
            lines = []
            for member, data, role_type in absent[:15]:
                icon = role_icons.get(role_type, '')
                excused = data.get('excused')
                reason = data.get('reason', '')
                
                if excused == 'У/П':
                    status = f"📝{icon} {member.display_name}"
                    if reason:
                        status += f" — {reason}"
                else:
                    status = f"❌{icon} {member.display_name}"
                
                lines.append(status)
            
            if len(absent) > 15:
                lines.append(f"... и ещё {len(absent) - 15}")
            
            embed.add_field(
                name=f"❌ Отсутствовали ({len(absent)})",
                value="\n".join(lines) or "—",
                inline=True
            )
        
        # Статистика
        total = len(present) + len(absent)
        present_pct = (len(present) / total * 100) if total > 0 else 0
        
        embed.add_field(
            name="📊 Статистика",
            value=f"**Всего:** {total}\n"
                  f"**Были:** {len(present)} ({present_pct:.0f}%)\n"
                  f"**Отсутствовали:** {len(absent)}",
            inline=False
        )
        
        embed.set_footer(text=f"Запросил: {ctx.author.display_name}")
        
        # View с редактированием
        sunday = bool(matching_schedule.get('choices'))
        view = KVReportView(self.bot, absent, date_str, guild_id, is_live=False,
                            sunday=sunday, sunday_choices=matching_schedule.get('choices'))

        await ctx.send(embed=embed, view=view)
    
    # ========================================
    # РЕДАКТИРОВАНИЕ !kvedit
    # ========================================
    
    @commands.command(name='kvedit', aliases=['kedit'])
    @officer_or_higher()
    async def kv_edit(self, ctx: commands.Context, date: str = None, member: discord.Member = None, *, status: str = None):
        """
        Редактировать посещаемость КВ
        !kvedit - меню выбора даты
        !kvedit 17-01-2026 - меню редактирования
        !kvedit 17-01-2026 @user присутствовал
        !kvedit 17-01-2026 @user отсутствовал
        !kvedit 17-01-2026 @user уп "Причина"
        """
        guild_id = ctx.guild.id
        
        if not date:
            # Показать список дат
            async with self.bot.db.execute('''
                SELECT DISTINCT date FROM kv_attendance 
                WHERE guild_id = ? 
                ORDER BY date DESC LIMIT 10
            ''', (guild_id,)) as cursor:
                dates = await cursor.fetchall()
            
            if not dates:
                await ctx.send("❌ Нет данных о КВ!")
                return
            
            embed = discord.Embed(
                title="📝 Редактирование КВ",
                description="Выберите дату:\n\n" + "\n".join([
                    f"• `!kvedit {format_date(date_str=d[0])}`" for d in dates
                ]),
                color=discord.Color.blue()
            )
            await ctx.send(embed=embed)
            return
        
        # Парсим дату
        target_date = parse_date(date)
        if not target_date:
            await ctx.send("❌ Неверный формат даты!")
            return
        
        date_str = date_for_db(target_date)
        
        if not member:
            # Показать меню редактирования
            embed = discord.Embed(
                title=f"📝 Редактирование КВ за {format_date(target_date)}",
                description="Используйте команду:\n"
                           "`!kvedit <дата> @user <статус>`\n\n"
                           "**Статусы:**\n"
                           "• `присутствовал` / `был`\n"
                           "• `отсутствовал` / `небыл`\n"
                           "• `уп` / `у/п` — уважительная причина",
                color=discord.Color.blue()
            )
            
            view = KVEditView(self.bot, date_str, guild_id)
            await ctx.send(embed=embed, view=view)
            return
        
        if not status:
            await ctx.send("❌ Укажите статус: `присутствовал`, `отсутствовал`, `уп`")
            return
        
        # Применяем статус
        status_lower = status.lower().split()[0]
        
        if status_lower in ['присутствовал', 'был', 'yes', 'да', '+', '1']:
            await self.bot.db.execute('''
                UPDATE kv_attendance 
                SET present = 1, excused = NULL, reason = 'Изменено вручную', processed_by = ?
                WHERE guild_id = ? AND date = ? AND user_id = ?
            ''', (ctx.author.id, guild_id, date_str, member.id))
            await self.bot.db.commit()
            await ctx.send(f"✅ **{member.display_name}** отмечен как **присутствующий**")
        
        elif status_lower in ['отсутствовал', 'небыл', 'no', 'нет', '-', '0']:
            await self.bot.db.execute('''
                UPDATE kv_attendance 
                SET present = 0, excused = NULL, reason = NULL, processed_by = ?
                WHERE guild_id = ? AND date = ? AND user_id = ?
            ''', (ctx.author.id, guild_id, date_str, member.id))
            await self.bot.db.commit()
            await ctx.send(f"❌ **{member.display_name}** отмечен как **отсутствующий**")
        
        elif status_lower in ['уп', 'у/п', 'excused']:
            # Парсим причину
            reason = "Не указана"
            parts = status.split(maxsplit=1)
            if len(parts) > 1:
                reason = parts[1].strip('"\'')
            
            await self.bot.db.execute('''
                UPDATE kv_attendance 
                SET present = 0, excused = 'У/П', reason = ?, processed_by = ?
                WHERE guild_id = ? AND date = ? AND user_id = ?
            ''', (reason, ctx.author.id, guild_id, date_str, member.id))
            await self.bot.db.commit()
            await ctx.send(f"📝 **{member.display_name}** — У/П ({reason})")
        
        else:
            await ctx.send("❌ Неизвестный статус! Используйте: `присутствовал`, `отсутствовал`, `уп`")
    
    # ========================================
    # ЛИЧНАЯ СТАТИСТИКА !me
    # ========================================
    
    @commands.command(name='me', aliases=['я', 'мой', 'mystats'])
    async def my_stats(self, ctx: commands.Context, member: discord.Member = None):
        """
        Личная статистика
        !me - своя статистика
        !me @user - статистика другого (офицеры+)
        """
        # Если запрашивают чужую статистику - проверяем права
        if member and member != ctx.author:
            if not ctx.author.guild_permissions.administrator:
                if not await self.bot.has_permission(ctx.author, 'officer'):
                    await ctx.send("❌ Только офицеры могут смотреть чужую статистику!")
                    return
        
        target = member or ctx.author
        guild_id = ctx.guild.id
        
        # КВ статистика
        async with self.bot.db.execute('''
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN present = 1 THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN excused = 'У/П' THEN 1 ELSE 0 END) as excused,
                SUM(vc_time_seconds) as total_time
            FROM kv_attendance
            WHERE guild_id = ? AND user_id = ?
        ''', (guild_id, target.id)) as cursor:
            kv_stats = await cursor.fetchone()

        # Голосовые сессии
        async with self.bot.db.execute('''
            SELECT COUNT(*), SUM(duration_seconds)
            FROM voice_sessions
            WHERE guild_id = ? AND user_id = ?
        ''', (guild_id, target.id)) as cursor:
            voice_stats = await cursor.fetchone()
        
        # Роль
        role_type = self.bot.get_member_role_type(target)
        role_icons = RANK_ICONS
        role_names = RANK_NAMES
        
        embed = discord.Embed(
            title=f"📊 Статистика: {target.display_name}",
            color=discord.Color.blue(),
            timestamp=datetime.now()
        )
        embed.set_thumbnail(url=target.display_avatar.url)
        
        embed.add_field(
            name="👤 Роль",
            value=f"{role_icons.get(role_type, '')} {role_names.get(role_type, role_type)}",
            inline=True
        )
        
        # КВ
        if kv_stats and kv_stats[0] > 0:
            total, present, excused, vc_time = kv_stats
            present = present or 0
            excused = excused or 0
            vc_time = vc_time or 0
            pct = (present / total * 100) if total > 0 else 0
            
            embed.add_field(
                name="⚔️ КВ",
                value=f"Посещено: **{present}/{total}** ({pct:.0f}%)\n"
                      f"У/П: **{excused}**\n"
                      f"Время: **{format_duration(vc_time)}**",
                inline=True
            )
            
            # Прогресс-бар
            bar_length = 10
            filled = int(bar_length * pct / 100)
            bar = "█" * filled + "░" * (bar_length - filled)
            embed.add_field(name="📈 КВ прогресс", value=f"`[{bar}]` {pct:.0f}%", inline=True)
        else:
            embed.add_field(name="⚔️ КВ", value="Нет данных", inline=True)

        # Голосовые
        if voice_stats and voice_stats[0] > 0:
            sessions, total_time = voice_stats
            total_time = total_time or 0
            embed.add_field(
                name="🎤 Голосовые",
                value=f"Сессий: **{sessions}**\n"
                      f"Время: **{format_duration(total_time)}**",
                inline=True
            )
        
        embed.set_footer(text=f"Запросил: {ctx.author.display_name}")
        
        await ctx.send(embed=embed)
    
    # ========================================
    # ЭКСПОРТ EXCEL
    # ========================================
    
    def _event_name_for_date(self, guild_id: int, date_str: str) -> str:
        """Название события КВ для даты (с учётом воскресного выбора)."""
        try:
            wd = datetime.strptime(date_str, '%Y-%m-%d').weekday()
        except ValueError:
            return 'КВ'
        for s in self.bot.get_guild_schedules(guild_id):
            if wd in s['days_of_week']:
                return self.bot.get_event_name(guild_id, date_str, s)
        return 'КВ'

    @commands.command(name='export', aliases=['экспорт', 'excel'])
    @officer_or_higher()
    async def export_excel(self, ctx: commands.Context, date: str = None):
        """
        Экспорт посещаемости КВ в Excel.
        Лист 1 — посещаемость за день (по умолчанию сегодня).
        Лист 2 — общая посещаемость по всем датам.
        !export            — за сегодня
        !export 17-01-2026 — лист дня за конкретную дату
        """
        if not OPENPYXL_AVAILABLE:
            await ctx.send("❌ openpyxl не установлен!")
            return

        msg = await ctx.send("📊 Генерация Excel...")

        guild = ctx.guild
        guild_id = guild.id
        settings = self.bot.guild_settings.get(guild_id, {})
        kv_vc_id = settings.get('kv_vc_channel_id')

        # Дата для листа дня
        if date:
            parsed = parse_date(date)
            if not parsed:
                await msg.edit(content="❌ Неверный формат даты! Используйте `DD-MM-YYYY`")
                return
            day_dt = parsed
        else:
            day_dt = datetime.now(self.bot.timezone)
        day_str = date_for_db(day_dt)
        day_display = format_date(day_dt)

        role_ru = RANK_NAMES
        role_order = {rank: i for i, rank in enumerate(RANK_ORDER)}

        try:
            wb = Workbook()

            # Стили
            thin = Side(style='thin', color="BFBFBF")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            title_font = Font(bold=True, size=14, color="FFFFFF")
            title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center")

            members = sorted(
                [m for m in guild.members if not m.bot],
                key=lambda m: (role_order.get(self.bot.get_member_role_type(m), 9), m.display_name.lower())
            )

            # ============ ЛИСТ 1: ПОСЕЩАЕМОСТЬ ЗА ДЕНЬ ============
            event_name = self._event_name_for_date(guild_id, day_str)
            stages = self.bot.get_event_stages(event_name)  # [(начало, конец), ...]
            ws = wb.active
            ws.title = "Сегодня"

            base_headers1 = ["Ник", "Звание", "Статус", "Вход", "Выход", "Сессии (вход–выход)", "Σ Время на КВ"]
            stage_headers = [f"Этап {i}\n{s}-{e}" for i, (s, e) in enumerate(stages, 1)]
            headers1 = base_headers1 + stage_headers
            n1 = len(headers1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n1)
            tcell = ws.cell(1, 1, f"Посещаемость КВ — {event_name} — {day_display}")
            tcell.font = title_font
            tcell.fill = title_fill
            tcell.alignment = center

            for col, h in enumerate(headers1, 1):
                c = ws.cell(2, col, h)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = center

            base_widths = [22, 12, 12, 16, 16, 34, 16]
            for i, w in enumerate(base_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            for i in range(len(stage_headers)):
                ws.column_dimensions[get_column_letter(len(base_widths) + 1 + i)].width = 13

            # Время начала КВ для этого дня
            kv_start_str = '20:00'
            for s in self.bot.get_guild_schedules(guild_id):
                if day_dt.weekday() in s['days_of_week']:
                    kv_start_str = s['start_time']
                    break
            kv_start = datetime.strptime(kv_start_str, '%H:%M').time()

            # Сессии за день в канале КВ
            sessions_by_user = {}
            if kv_vc_id:
                async with self.bot.db.execute('''
                    SELECT user_id, join_time, leave_time, duration_seconds
                    FROM voice_sessions
                    WHERE guild_id = ? AND date = ? AND channel_id = ?
                    ORDER BY join_time
                ''', (guild_id, day_str, kv_vc_id)) as cursor:
                    async for row in cursor:
                        sessions_by_user.setdefault(row[0], []).append((row[1], row[2], row[3] or 0))

            # Отметки посещаемости за день
            day_att = {}
            async with self.bot.db.execute('''
                SELECT user_id, present, excused FROM kv_attendance
                WHERE guild_id = ? AND date = ?
            ''', (guild_id, day_str)) as cursor:
                async for row in cursor:
                    day_att[row[0]] = (row[1], row[2])

            # Присутствие на этапе = сессия пересекается с окном этапа
            day_date = day_dt.date()

            def stage_present(sess, start_str, end_str):
                sst = datetime.combine(day_date, datetime.strptime(start_str, '%H:%M').time())
                sen = datetime.combine(day_date, datetime.strptime(end_str, '%H:%M').time())
                for j, l, d in sess:
                    js = datetime.fromisoformat(j).replace(tzinfo=None)
                    le = datetime.fromisoformat(l).replace(tzinfo=None) if l else sen
                    if js < sen and le > sst:
                        return True
                return False

            r = 3
            for m in members:
                sessions = sessions_by_user.get(m.id, [])
                total_dur = sum(s[2] for s in sessions)
                att = day_att.get(m.id)
                excused = att[1] if att else None
                present = bool(att[0]) if att else False
                if sessions:
                    present = True

                if sessions:
                    fj = datetime.fromisoformat(sessions[0][0])
                    entry = f"был до {kv_start_str}" if fj.time() < kv_start else fj.strftime('%H:%M')
                    last_leave = sessions[-1][1]
                    exit_ = datetime.fromisoformat(last_leave).strftime('%H:%M') if last_leave else "ещё в ГС"
                else:
                    entry = "—"
                    exit_ = "—"

                parts = []
                for j, l, d in sessions:
                    js = datetime.fromisoformat(j).strftime('%H:%M')
                    ls = datetime.fromisoformat(l).strftime('%H:%M') if l else "…"
                    parts.append(f"{js}–{ls}")
                sessions_str = ", ".join(parts) if parts else "—"

                if excused == 'У/П':
                    status, fill = "У/П", yellow_fill
                elif present:
                    status, fill = "✅ был", green_fill
                else:
                    status, fill = "❌ не был", red_fill

                role_type = self.bot.get_member_role_type(m)
                values = [
                    m.display_name,
                    role_ru.get(role_type, role_type),
                    status,
                    entry,
                    exit_,
                    sessions_str,
                    format_duration(total_dur) if total_dur > 0 else "—",
                ]
                for col, v in enumerate(values, 1):
                    c = ws.cell(r, col, v)
                    c.border = border
                    c.fill = fill
                    c.alignment = left_align if col in (1, 6) else center

                # Этапы: присутствие на каждом
                for k, (s_start, s_end) in enumerate(stages):
                    col = len(base_headers1) + 1 + k
                    if excused == 'У/П':
                        sym, sfill = "У/П", yellow_fill
                    elif sessions:
                        if stage_present(sessions, s_start, s_end):
                            sym, sfill = "✅", green_fill
                        else:
                            sym, sfill = "☐", red_fill
                    elif present:
                        sym, sfill = "—", gray_fill
                    else:
                        sym, sfill = "☐", red_fill
                    c = ws.cell(r, col, sym)
                    c.border = border
                    c.fill = sfill
                    c.alignment = center
                r += 1

            ws.freeze_panes = "A3"
            ws.row_dimensions[2].height = 28

            # ============ ЛИСТ 2: ОБЩАЯ ПОСЕЩАЕМОСТЬ ============
            ws2 = wb.create_sheet("Посещаемость")

            async with self.bot.db.execute('''
                SELECT DISTINCT date FROM kv_attendance WHERE guild_id = ? ORDER BY date
            ''', (guild_id,)) as cursor:
                all_dates = [row[0] for row in await cursor.fetchall()]

            att_map = {}
            async with self.bot.db.execute('''
                SELECT user_id, date, present, excused FROM kv_attendance WHERE guild_id = ?
            ''', (guild_id,)) as cursor:
                async for row in cursor:
                    att_map[(row[0], row[1])] = (row[2], row[3])

            abbr = {'Потасовка': 'Пот', 'Турнир': 'Тур', 'Захват базы': 'Зхв'}
            total_cols = max(2 + len(all_dates) + 2, 2)

            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            t2 = ws2.cell(1, 1, "Общая посещаемость КВ — StalZone")
            t2.font = title_font
            t2.fill = title_fill
            t2.alignment = center

            for col, h in enumerate(["Ник", "Звание"], 1):
                c = ws2.cell(2, col, h)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = center

            for i, d in enumerate(all_dates):
                col = 3 + i
                try:
                    dd = datetime.strptime(d, '%Y-%m-%d')
                    ev = self._event_name_for_date(guild_id, d)
                    label = f"{dd.strftime('%d.%m')}\n{abbr.get(ev, ev[:3])}"
                except ValueError:
                    label = d
                c = ws2.cell(2, col, label)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = center
                ws2.column_dimensions[get_column_letter(col)].width = 7

            sum_col = 3 + len(all_dates)
            pct_col = sum_col + 1
            for col, h in ((sum_col, "Был"), (pct_col, "%")):
                c = ws2.cell(2, col, h)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = center
                ws2.column_dimensions[get_column_letter(col)].width = 8

            ws2.column_dimensions['A'].width = 22
            ws2.column_dimensions['B'].width = 12

            r = 3
            for m in members:
                role_type = self.bot.get_member_role_type(m)
                cn = ws2.cell(r, 1, m.display_name)
                cn.border = border
                cn.alignment = left_align
                cr = ws2.cell(r, 2, role_ru.get(role_type, role_type))
                cr.border = border
                cr.alignment = center

                present_count = 0
                counted = 0
                for i, d in enumerate(all_dates):
                    col = 3 + i
                    rec = att_map.get((m.id, d))
                    if rec is None:
                        sym, fill = "–", gray_fill
                    else:
                        pres, exc = rec
                        if exc == 'У/П':
                            sym, fill = "У/П", yellow_fill
                        elif pres:
                            sym, fill = "✅", green_fill
                            present_count += 1
                            counted += 1
                        else:
                            sym, fill = "☐", red_fill
                            counted += 1
                    c = ws2.cell(r, col, sym)
                    c.fill = fill
                    c.border = border
                    c.alignment = center

                sc = ws2.cell(r, sum_col, present_count)
                sc.border = border
                sc.alignment = center
                pct = (present_count / counted * 100) if counted else 0
                pc = ws2.cell(r, pct_col, f"{pct:.0f}%")
                pc.border = border
                pc.alignment = center
                r += 1

            # Легенда
            r += 1
            lc = ws2.cell(r, 1, "Легенда:")
            lc.font = Font(bold=True)
            for i, (txt, fl) in enumerate((("✅ был", green_fill), ("☐ не был", red_fill),
                                           ("У/П уваж.", yellow_fill), ("– не было КВ", gray_fill))):
                c = ws2.cell(r, 2 + i, txt)
                c.fill = fl
                c.border = border
                c.alignment = center

            ws2.freeze_panes = "C3"
            ws2.row_dimensions[2].height = 28

            # Сохранение
            excel_path = self.bot.config.get('EXCEL_PATH', 'data/clan_attendance.xlsx')
            wb.save(excel_path)

            await msg.delete()

            embed = discord.Embed(
                title="📊 Excel отчёт готов!",
                color=discord.Color.green()
            )
            embed.add_field(
                name="📋 Листы",
                value=(f"• **Сегодня** — посещаемость за {day_display} ({event_name})\n"
                       f"• **Посещаемость** — все даты: {len(all_dates)} КВ, {len(members)} участников"),
                inline=False
            )
            await ctx.send(embed=embed, file=discord.File(excel_path))

        except Exception as e:
            logger.error(f"Ошибка экспорта: {e}")
            import traceback
            traceback.print_exc()
            await msg.edit(content=f"❌ Ошибка: {e}")


    # ========================================
    # СТАТИСТИКА
    # ========================================
    
    @commands.command(name='stats', aliases=['статистика'])
    async def stats(self, ctx: commands.Context, member: discord.Member = None):
        """Статистика пользователя"""
        await ctx.invoke(self.bot.get_command('me'), member=member)
    
    @commands.command(name='top10', aliases=['топ10', 'топ'])
    async def top10(self, ctx: commands.Context):
        """Топ-10 по посещаемости КВ"""
        guild_id = ctx.guild.id
        
        async with self.bot.db.execute('''
            SELECT user_id, discord_name,
                   COUNT(*) as total,
                   SUM(CASE WHEN present = 1 THEN 1 ELSE 0 END) as present,
                   SUM(vc_time_seconds) as vc_time
            FROM kv_attendance
            WHERE guild_id = ?
            GROUP BY user_id
            ORDER BY present DESC, vc_time DESC
            LIMIT 10
        ''', (guild_id,)) as cursor:
            rows = await cursor.fetchall()
        
        if not rows:
            await ctx.send("📊 Нет данных")
            return
        
        embed = discord.Embed(
            title="🏆 Топ-10 по КВ",
            color=discord.Color.gold()
        )
        
        medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
        
        lines = []
        for i, (user_id, name, total, present, vc_time) in enumerate(rows):
            vc_time = vc_time or 0
            pct = (present / total * 100) if total > 0 else 0
            lines.append(f"{medals[i]} **{name}** — {present}/{total} ({pct:.0f}%) • {format_duration(vc_time)}")
        
        embed.description = "\n".join(lines)
        await ctx.send(embed=embed)
    
    @commands.command(name='online', aliases=['онлайн'])
    async def online(self, ctx: commands.Context):
        """Кто сейчас в голосовых каналах"""
        guild = ctx.guild
        
        embed = discord.Embed(
            title="🎤 Онлайн в голосовых",
            color=discord.Color.green()
        )
        
        total = 0
        for vc in guild.voice_channels:
            members = [m for m in vc.members if not m.bot]
            if members:
                total += len(members)
                names = ", ".join([m.display_name for m in members[:10]])
                if len(members) > 10:
                    names += f" и ещё {len(members)-10}"
                embed.add_field(name=f"🔊 {vc.name} ({len(members)})", value=names, inline=False)
        
        if total == 0:
            embed.description = "Все каналы пусты"
        else:
            embed.description = f"**Всего:** {total}"
        
        await ctx.send(embed=embed)


async def setup(bot):
    await bot.add_cog(AttendanceCog(bot))
